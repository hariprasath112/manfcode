from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import time
from openpyxl import load_workbook
from tkinter.filedialog import askopenfilename

# Initialize WebDriver
username = input("Enter username: ")
password = input("Enter password: ")
column = input("Enter a column containing product codes: ")
rcolumn = input("Enter a column to write manufacturer codes: ")
filename = askopenfilename()

# Load Excel workbook and select the active sheet
wb = load_workbook(filename)
ws = wb.active

driver = webdriver.Chrome()
driver.maximize_window()
driver.get("https://enutrition.sysco.com/eNutrition/Login/Index/")

def login():
    """Login to the website."""
    try:
        userBox = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="userid"]')))
        userBox.send_keys(username)

        passBox = driver.find_element(By.XPATH, '//*[@id="pwd"]')
        passBox.send_keys(password)

        loginButton = driver.find_element(By.XPATH, '/html/body/div[1]/form/div[2]/div[1]/div[1]/table/tbody/tr[4]/td/input')
        loginButton.click()

        # Wait until the next page loads and click on the link
        clickLink = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/form/div[2]/div[2]/table/tbody/tr/td/table/tbody/tr[4]/td/table/tbody/tr[1]/td/table/tbody/tr[2]/td/table/tbody/tr/td/a')))
        clickLink.click()

    except TimeoutException:
        pass

def get_mpc_from_web(code):
    """Retrieve MPC code from the website using the provided product code."""
    try:
        # If the search page selector exists, click it
        try:
            searchPage = WebDriverWait(driver, 1).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="cc_NutritionItemSelector"]')))
            searchPage.click()
        except TimeoutException:
            pass  # Continue if already on the search page

        # Wait until the search bar is clickable, clear it, and enter the code
        searchBar = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="t4_qquery2"]')))
        searchBar.clear()
        searchBar.send_keys(code)

        # Add a small delay before clicking the search button
        time.sleep(1)
        searchButton = WebDriverWait(driver, 1).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="t4search"]')))
        searchButton.click()

        # Click error OK if it appears
        try:
            alertOK = WebDriverWait(driver, 1).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="alertOK"]')))
            alertOK.click()
        except TimeoutException:
            pass

        # Wait for search results and click the first result
        try:
            firstResult = WebDriverWait(driver, 4).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="hrt4_cG3_0_0"]')))
            firstResult.click()
        except TimeoutException:
            return None

        # Extract the result information (MPC code)
        try:
            productInfo = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, '/html/body/form/div[2]/div[2]/fieldset/table/tbody/tr[2]/td[1]/table/tbody/tr[3]/td')))
            return productInfo.text  # Assuming this is the MPC code
        except TimeoutException:
            return None

    except Exception:
        return None

def process_excel():
    """Process the Excel file, retrieve MPC codes, and write them back into the file."""
    # Loop through each row in column containing Product Codes
    for row in range(2, ws.max_row + 1):  # Start from the second row (skip headers)
        product_code = ws[f"{column}{row}"].value
        if product_code:
            mpc_code = get_mpc_from_web(product_code)

            if mpc_code:
                ws[f"{rcolumn}{row}"] = mpc_code.split(": ")[1]  # Write the MPC code into the designated column
            else:
                ws[f"{rcolumn}{row}"] = "N/A"  # Mark as N/A if MPC not found

    # Save the updated workbook
    wb.save(filename)
    print("Excel file updated successfully.")

# Run the script
login()
process_excel()
